"""Excel export for empty and styled reports."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from app.excel_format import ExcelReportPayload, build_styled_workbook
from app.report_excel import build_full_excel_from_report, build_top10_excel_from_report


EMPTY_LIVE_REPORT = {
    "summary_text": "Live-поток: пока нет обращений граждан",
    "top3": [],
    "top10": [],
    "all": [],
    "topics": [],
    "groups": [],
    "reasons": [],
}


SAMPLE_REPORT = {
    "summary_text": "За период зафиксирован рост обращений по ЖКХ в трёх муниципалитетах.",
    "stats": {
        "start_date": "2025-01-01",
        "end_date": "2025-03-31",
        "rows_processed": 1200,
        "problem_count": 890,
    },
    "top3": [
        {
            "rank": 1,
            "муниципалитет": "г. Омск",
            "score": 82,
            "total_incidents": 500,
            "problem_count": 420,
            "critical_count": 12,
            "топ_тема": "ЖКХ",
            "summary_text": "Преобладают жалобы на отопление.",
        },
        {
            "rank": 2,
            "муниципалитет": "Омский район",
            "score": 71,
            "total_incidents": 200,
            "problem_count": 150,
            "critical_count": 4,
            "топ_тема": "Дороги",
        },
        {
            "rank": 3,
            "муниципалитет": "Исilkul",
            "score": 65,
            "total_incidents": 120,
            "problem_count": 90,
            "critical_count": 1,
            "топ_тема": "Свет",
        },
    ],
    "top10": [
        {
            "rank": i + 1,
            "муниципалитет": name,
            "score": 90 - i * 5,
            "total_incidents": 300 - i * 20,
            "problem_count": 250 - i * 15,
            "critical_count": max(0, 10 - i),
            "топ_тема": "ЖКХ" if i % 2 == 0 else "Дороги",
        }
        for i, name in enumerate(
            ["г. Омск", "Омский район", "Исilkul", "Калачинский", "Тарский", "Москаленский",
             "Павлоградский", "Нижнеомский", "Кормиловский", "Черлакский"]
        )
    ],
    "all": [],
    "topics": [
        {"муниципалитет": "г. Омск", "тема": "Отопление", "count": 120},
        {"муниципалитет": "г. Омск", "тема": "Мусор", "count": 80},
        {"муниципалитет": "Омский район", "тема": "Дороги", "count": 45},
    ],
    "groups": [
        {"муниципалитет": "г. Омск", "группа": "ЖКХ", "count": 200},
        {"муниципалитет": "г. Омск", "группа": "Благоустройство", "count": 90},
    ],
    "reasons": [
        {
            "муниципалитет": "г. Омск",
            "топ_тема": "Отопление",
            "ключевые_темы": "Отопление(120); Мусор(80)",
            "summary_text": "Жалобы на низкую температуру в домах.",
        }
    ],
    "severity_breakdown": [
        {"муниципалитет": "г. Омск", "severity": 3, "label": "Высокая", "count": 100},
        {"муниципалитет": "г. Омск", "severity": 4, "label": "Критическая", "count": 12},
        {"муниципалитет": "г. Омск", "severity": 2, "label": "Средняя", "count": 200},
    ],
    "resolved_stats": [
        {
            "муниципалитет": "г. Омск",
            "problem_count": 420,
            "resolved_count": 210,
            "resolved_pct": 50.0,
        },
        {
            "муниципалитет": "Омский район",
            "problem_count": 150,
            "resolved_count": 60,
            "resolved_pct": 40.0,
        },
    ],
}


@pytest.mark.parametrize(
    "report",
    [
        EMPTY_LIVE_REPORT,
        {},
        {"top3": None, "top10": None, "all": None},
    ],
)
def test_build_full_excel_from_empty_report(tmp_path: Path, report: dict):
    path = build_full_excel_from_report(report, tmp_path)
    assert path.exists()
    assert path.stat().st_size > 0

    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) >= 1
    assert wb.active.title == "Сводка"


def test_build_top10_excel_from_empty_report(tmp_path: Path):
    path = build_top10_excel_from_report(EMPTY_LIVE_REPORT, tmp_path)
    assert path.exists()
    assert path.stat().st_size > 0

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Сводка"]
    assert "Live-поток" in str(wb["Сводка"]["A3"].value or wb["Сводка"]["A4"].value or "")


def test_styled_full_excel_has_summary_and_charts(tmp_path: Path):
    path = build_full_excel_from_report(SAMPLE_REPORT, tmp_path)
    wb = openpyxl.load_workbook(path)

    assert "Сводка" in wb.sheetnames
    assert "Top-10" in wb.sheetnames
    assert "Top-3" in wb.sheetnames
    assert "Графики" in wb.sheetnames
    assert "Темы" in wb.sheetnames

    summary = wb["Сводка"]
    assert summary["A1"].value and "ZeroProblems" in summary["A1"].value

    charts_ws = wb["Графики"]
    assert len(charts_ws._charts) >= 1

    top10 = wb["Top-10"]
    assert top10["A3"].font.bold is True  # header row styled


def test_excel_includes_resolved_markup(tmp_path: Path):
    path = build_full_excel_from_report(SAMPLE_REPORT, tmp_path)
    wb = openpyxl.load_workbook(path)

    assert "Решённость" in wb.sheetnames
    resolved = wb["Решённость"]
    headers = [resolved.cell(row=3, column=c).value for c in range(1, 6)]
    assert "Муниципалитет" in headers
    assert "Решено" in headers
    assert "Нерешённых" in headers
    assert "Доля решённых" in headers

    top10 = wb["Top-10"]
    top_headers = [top10.cell(row=3, column=c).value for c in range(1, 12)]
    assert "Решено" in top_headers
    assert "Нерешённых" in top_headers
    assert "Доля решённых" in top_headers

    summary = wb["Сводка"]
    kpi_labels = {summary.cell(row=4, column=c).value for c in range(1, 8)}
    assert "Решено" in kpi_labels
    assert "Нерешённых" in kpi_labels


def test_excel_labeled_sample_has_status_column():
    sample = pd.DataFrame(
        {
            "муниципалитет": ["А", "А"],
            "текст": ["жалоба 1", "жалоба 2"],
            "severity": [2, 3],
            "итог": ["Решено", "Отклонено"],
        }
    )
    wb = build_styled_workbook(
        ExcelReportPayload(
            summary_text="test",
            top10=pd.DataFrame([{"rank": 1, "муниципалитет": "А", "score": 50, "problem_count": 2}]),
            labeled_sample=sample,
            include_labeled_sample=True,
            include_all_municipalities=False,
        )
    )
    ws = wb["Примеры"]
    headers = [ws.cell(row=3, column=c).value for c in range(1, 6)]
    assert headers[0] == "Статус"
    assert ws.cell(row=4, column=1).value == "Решено"
    assert ws.cell(row=5, column=1).value == "Нерешено"


def test_styled_top10_excel_has_summary_and_charts(tmp_path: Path):
    path = build_top10_excel_from_report(SAMPLE_REPORT, tmp_path)
    wb = openpyxl.load_workbook(path)

    assert "Сводка" in wb.sheetnames
    assert "Top-10" in wb.sheetnames
    assert "Графики" in wb.sheetnames
    assert "Все МО" not in wb.sheetnames

    charts_ws = wb["Графики"]
    assert len(charts_ws._charts) >= 1

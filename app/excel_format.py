"""Оформление Excel-отчётов ZeroProblems: стили, сводка, графики."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.resolved import is_resolved_row

# ── Палитра ─────────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor="B91C1C")
FILL_TITLE = PatternFill("solid", fgColor="1E293B")
FILL_KPI = PatternFill("solid", fgColor="FEF2F2")
FILL_ALT = PatternFill("solid", fgColor="F8FAFC")
FILL_CRITICAL = PatternFill("solid", fgColor="FEE2E2")
FILL_WARN = PatternFill("solid", fgColor="FFEDD5")
FILL_OK = PatternFill("solid", fgColor="ECFDF5")

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Calibri", size=11, bold=True, color="1E293B")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Calibri", size=10, color="1E293B")
FONT_KPI_VAL = Font(name="Calibri", size=14, bold=True, color="B91C1C")
FONT_KPI_LBL = Font(name="Calibri", size=9, color="64748B")

THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Данные для диаграмм — в правой части листа «Графики» (скрытые листы ломают подписи в Excel).
_CHART_DATA_START_COL = 11
_CHART_ROW_HEIGHT_CM = 0.53
_CHART_LAYOUT_ATTR = "_zp_chart_layout"


@dataclass
class _ChartLayout:
    ws: Any
    next_row: int
    data_col: int = _CHART_DATA_START_COL

    def write_table(self, headers: list[str], rows: list[tuple]) -> tuple[int, int, int]:
        """Возвращает (data_start, data_end, start_col) для Reference."""
        start_col = self.data_col
        header_row = 1
        for i, header in enumerate(headers):
            cell = self.ws.cell(row=header_row, column=start_col + i, value=header)
            cell.font = Font(name="Calibri", size=9, bold=True, color="64748B")
        data_start = header_row + 1
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                self.ws.cell(row=data_start + ri, column=start_col + ci, value=val)
        data_end = data_start + max(len(rows) - 1, 0)
        self.data_col += len(headers) + 2
        return data_start, data_end, start_col

    def add(self, chart, col: str = "A") -> None:
        anchor = f"{col}{self.next_row}"
        self.ws.add_chart(chart, anchor)
        span = max(int(chart.height / _CHART_ROW_HEIGHT_CM) + 5, 24)
        self.next_row += span


def _apply_bar_value_labels(chart) -> None:
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showLegendKey = False


def _configure_hbar_chart(chart, *, value_title: str, n_bars: int) -> None:
    chart.type = "bar"
    chart.legend = None
    chart.style = 10
    chart.x_axis.title = value_title
    chart.y_axis.title = None
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.height = max(10, min(20, 4 + n_bars * 1.15))
    chart.width = 24
    _apply_bar_value_labels(chart)


def _configure_col_chart(chart, *, value_title: str, n_bars: int) -> None:
    chart.type = "col"
    chart.grouping = "clustered"
    chart.style = 10
    chart.y_axis.title = value_title
    chart.x_axis.title = None
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.height = max(10, min(18, 4 + n_bars * 0.9))
    chart.width = 24


def _chart_categories(ws, start_col: int, data_start: int, data_end: int) -> Reference:
    return Reference(ws, min_col=start_col, min_row=data_start, max_row=data_end)


def _chart_series(ws, start_col: int, data_start: int, data_end: int) -> Reference:
    return Reference(ws, min_col=start_col, min_row=data_start - 1, max_row=data_end)


def _ensure_charts_layout(wb: Workbook, subtitle: str = "") -> _ChartLayout:
    cached = getattr(wb, _CHART_LAYOUT_ATTR, None)
    if cached is not None:
        return cached

    if "Графики" in wb.sheetnames:
        ws = wb["Графики"]
        start_row = max(ws.max_row + 3, 4)
    else:
        ws = wb.create_sheet("Графики")
        start_row = _write_title_block(ws, "Визуализация", subtitle) + 2

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    layout = _ChartLayout(ws=ws, next_row=start_row)
    setattr(wb, _CHART_LAYOUT_ATTR, layout)
    return layout

COLUMN_LABELS: dict[str, str] = {
    "rank": "Место",
    "district_id": "№",
    "муниципалитет": "Муниципалитет",
    "score": "Индекс проблемности",
    "health_score": "Индекс проблемности",
    "rating_score": "Рейтинг",
    "total_incidents": "Всего обращений",
    "problem_count": "Проблемных",
    "critical_count": "Критических (кл.4)",
    "high_count": "Высоких (кл.3+)",
    "problem_share": "Доля проблемных",
    "severity_mean": "Средний класс",
    "severity_p90": "Класс P90",
    "severity_sum": "Сумма классов",
    "топ_тема": "Главная тема",
    "топ_группа": "Главная группа",
    "ключевые_темы": "Ключевые темы",
    "summary_text": "Аналитическая справка",
    "summary_paragraph": "Справка (развёрнутая)",
    "summary": "Справка",
    "тема": "Тема",
    "группа": "Группа тем",
    "count": "Количество",
    "severity": "Класс",
    "label": "Уровень",
    "причина": "Причина / проблема",
    "описание": "Описание",
    "resolved_count": "Решено",
    "open_count": "Нерешённых",
    "resolved_pct": "Доля решённых",
    "статус": "Статус",
    "итог": "Итог",
}

RANKING_PREFERRED = [
    "rank",
    "муниципалитет",
    "score",
    "total_incidents",
    "problem_count",
    "resolved_count",
    "open_count",
    "resolved_pct",
    "critical_count",
    "high_count",
    "problem_share",
    "severity_mean",
    "топ_тема",
    "топ_группа",
    "ключевые_темы",
    "summary_text",
    "summary_paragraph",
]

REASONS_PREFERRED = [
    "муниципалитет",
    "причина",
    "топ_тема",
    "топ_группа",
    "ключевые_темы",
    "problem_count",
    "summary_text",
]

TOPIC_PREFERRED = ["муниципалитет", "тема", "count"]
GROUP_PREFERRED = ["муниципалитет", "группа", "count"]
RESOLVED_STATS_PREFERRED = [
    "муниципалитет",
    "problem_count",
    "resolved_count",
    "open_count",
    "resolved_pct",
]


@dataclass
class ExcelReportPayload:
    summary_text: str = ""
    stats: dict[str, Any] = field(default_factory=dict)
    top_all: pd.DataFrame = field(default_factory=pd.DataFrame)
    top10: pd.DataFrame = field(default_factory=pd.DataFrame)
    top3: pd.DataFrame = field(default_factory=pd.DataFrame)
    topics_df: pd.DataFrame = field(default_factory=pd.DataFrame)
    groups_df: pd.DataFrame = field(default_factory=pd.DataFrame)
    reasons_df: pd.DataFrame = field(default_factory=pd.DataFrame)
    severity_breakdown: list[dict] = field(default_factory=list)
    muni_summaries: pd.DataFrame | None = None
    top3_summaries: pd.DataFrame | None = None
    labeled_sample: pd.DataFrame | None = None
    meta_rows: list[tuple[str, str]] | None = None
    include_all_municipalities: bool = True
    include_labeled_sample: bool = False
    resolved_stats: list[dict] = field(default_factory=list)


def _resolved_stats_map(resolved_stats: list[dict]) -> dict[str, dict]:
    return {str(row.get("муниципалитет", "")).strip(): row for row in resolved_stats}


def _enrich_ranking_df(df: pd.DataFrame, resolved_stats: list[dict]) -> pd.DataFrame:
    """Добавляет resolved_count / open_count / resolved_pct по МО (как в PDF)."""
    if df.empty or not resolved_stats:
        return df
    stats_map = _resolved_stats_map(resolved_stats)
    out = df.copy()
    resolved_counts: list[int | None] = []
    open_counts: list[int | None] = []
    resolved_pcts: list[float | None] = []
    for _, row in out.iterrows():
        st = stats_map.get(str(row.get("муниципалитет", "")).strip())
        if not st:
            resolved_counts.append(None)
            open_counts.append(None)
            resolved_pcts.append(None)
            continue
        pc = int(st.get("problem_count") or 0)
        rc = int(st.get("resolved_count") or 0)
        resolved_counts.append(rc)
        open_counts.append(max(0, pc - rc))
        resolved_pcts.append(float(st.get("resolved_pct") or (round(100.0 * rc / pc, 1) if pc else 0.0)))
    out["resolved_count"] = resolved_counts
    out["open_count"] = open_counts
    out["resolved_pct"] = resolved_pcts
    return out


def _resolved_stats_df(resolved_stats: list[dict]) -> pd.DataFrame:
    if not resolved_stats:
        return pd.DataFrame()
    rows: list[dict] = []
    for row in resolved_stats:
        pc = int(row.get("problem_count") or 0)
        rc = int(row.get("resolved_count") or 0)
        rows.append(
            {
                "муниципалитет": str(row.get("муниципалитет", "")),
                "problem_count": pc,
                "resolved_count": rc,
                "open_count": max(0, pc - rc),
                "resolved_pct": float(row.get("resolved_pct") or (round(100.0 * rc / pc, 1) if pc else 0.0)),
            }
        )
    return pd.DataFrame(rows)


def _region_resolved_kpi(resolved_stats: list[dict]) -> tuple[int, int, float] | None:
    if not resolved_stats:
        return None
    total_problem = sum(int(r.get("problem_count") or 0) for r in resolved_stats)
    if total_problem <= 0:
        return None
    total_resolved = sum(int(r.get("resolved_count") or 0) for r in resolved_stats)
    total_resolved = max(0, min(total_problem, total_resolved))
    pct = round(100.0 * total_resolved / total_problem, 1)
    return total_resolved, total_problem - total_resolved, pct


def _add_resolution_status(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    status = out.apply(lambda row: "Решено" if is_resolved_row(row) else "Нерешено", axis=1)
    if "статус" in out.columns:
        out["статус"] = status
    else:
        out.insert(0, "статус", status)
    return out


def _label(col: str) -> str:
    return COLUMN_LABELS.get(col, col.replace("_", " ").capitalize())


def _pick_columns(df: pd.DataFrame, preferred: list[str]) -> list[str]:
    if df.empty:
        return []
    cols = [c for c in preferred if c in df.columns]
    if cols:
        return cols
    return list(df.columns)


def _prepare_df(df: pd.DataFrame, preferred: list[str]) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    for col in out.select_dtypes(include="object").columns:
        out[col] = out[col].fillna("").astype(str)
    cols = _pick_columns(out, preferred)
    out = out[cols]
    out.columns = [_label(c) for c in cols]
    if "Доля проблемных" in out.columns:
        out["Доля проблемных"] = pd.to_numeric(out["Доля проблемных"], errors="coerce").map(
            lambda x: f"{x * 100:.1f}%" if pd.notna(x) else ""
        )
    if "Доля решённых" in out.columns:
        out["Доля решённых"] = pd.to_numeric(out["Доля решённых"], errors="coerce").map(
            lambda x: f"{x:.1f}%" if pd.notna(x) else ""
        )
    return out


def _style_range(ws, row: int, col_start: int, col_end: int, fill=None, font=None, alignment=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        cell.border = BORDER


def _autofit(ws, min_w: int = 10, max_w: int = 48):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max((len(str(c.value or "")) for c in col_cells), default=min_w)
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def _write_title_block(ws, title: str, subtitle: str = "", start_row: int = 1) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 28
    row = start_row + 1
    if subtitle:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        sub = ws.cell(row=row, column=1, value=subtitle)
        sub.font = Font(name="Calibri", size=10, color="64748B")
        sub.alignment = Alignment(wrap_text=True)
        row += 1
    return row


def _write_kpi_row(ws, row: int, items: list[tuple[str, str]]) -> int:
    col = 1
    for label, value in items:
        vcell = ws.cell(row=row, column=col, value=value)
        vcell.font = FONT_KPI_VAL
        vcell.fill = FILL_KPI
        vcell.alignment = Alignment(horizontal="center", vertical="bottom")
        lcell = ws.cell(row=row + 1, column=col, value=label)
        lcell.font = FONT_KPI_LBL
        lcell.fill = FILL_KPI
        lcell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 18
        col += 1
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 28
    return row + 3


def write_summary_sheet(wb: Workbook, payload: ExcelReportPayload) -> None:
    ws = wb.active
    ws.title = "Сводка"
    stats = payload.stats or {}
    period = ""
    if stats.get("start_date") or stats.get("end_date"):
        period = f"Период: {stats.get('start_date', '—')} — {stats.get('end_date', '—')}"

    row = _write_title_block(
        ws,
        "ZeroProblems — аналитический отчёт",
        f"Омская область · {period} · сформирован {datetime.now().strftime('%d.%m.%Y %H:%M')}",
    )

    total = stats.get("rows_processed") or stats.get("total_incidents")
    if total is None and not payload.top_all.empty and "total_incidents" in payload.top_all.columns:
        total = int(payload.top_all["total_incidents"].sum())
    problems = stats.get("problem_count")
    if problems is None and not payload.top_all.empty and "problem_count" in payload.top_all.columns:
        problems = int(payload.top_all["problem_count"].sum())
    muni_count = len(payload.top_all) if not payload.top_all.empty else len(payload.top10)

    kpi_items = [
        ("Всего обращений", f"{int(total or 0):,}".replace(",", " ")),
        ("Проблемных", f"{int(problems or 0):,}".replace(",", " ")),
        ("Муниципалитетов", str(muni_count)),
        ("В Top-10", str(len(payload.top10))),
    ]
    region_resolved = _region_resolved_kpi(payload.resolved_stats)
    if region_resolved is not None:
        resolved_n, open_n, resolved_pct = region_resolved
        kpi_items = [
            ("Всего обращений", f"{int(total or 0):,}".replace(",", " ")),
            ("Проблемных", f"{int(problems or 0):,}".replace(",", " ")),
            ("Решено", f"{resolved_n:,}".replace(",", " ")),
            ("Нерешённых", f"{open_n:,}".replace(",", " ")),
            ("Доля решённых", f"{resolved_pct:.1f}%"),
        ]
    row = _write_kpi_row(ws, row, kpi_items)

    ws.cell(row=row, column=1, value="Краткая сводка").font = FONT_SUBTITLE
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 4, end_column=6)
    summary = (payload.summary_text or "Нет данных").strip()
    scell = ws.cell(row=row, column=1, value=summary)
    scell.font = FONT_BODY
    scell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 6

    if not payload.top3.empty:
        ws.cell(row=row, column=1, value="Top-3 критичных муниципалитетов").font = FONT_SUBTITLE
        row += 1
        for _, r in payload.top3.head(3).iterrows():
            muni = str(r.get("муниципалитет", ""))
            score = r.get("score", r.get("health_score", ""))
            prob = r.get("problem_count", "")
            theme = str(r.get("топ_тема", r.get("ключевые_темы", ""))).split(";")[0][:80]
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            line = f"• {muni} — индекс {score}, проблемных {prob}. {theme}"
            c = ws.cell(row=row, column=1, value=line)
            c.font = FONT_BODY
            c.alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    ws.cell(row=row, column=1, value="Структура книги").font = FONT_SUBTITLE
    row += 1
    sheets_hint = [
        "«Top-10» — рейтинг проблемных МО с графиком",
        "«Top-3» — наиболее критичные территории",
    ]
    if payload.include_all_municipalities:
        sheets_hint.append("«Все МО» — полный рейтинг")
    sheets_hint += [
        "«Темы» / «Группы» — структура обращений (нерешённые)",
        "«Причины» — ключевые проблемы по МО",
    ]
    if payload.resolved_stats:
        sheets_hint.append("«Решённость» — решено / нерешено по МО")
    sheets_hint.append("«Графики» — визуализация")
    if payload.include_labeled_sample:
        sheets_hint.append("«Примеры» — размеченные обращения со статусом решённости")
    for hint in sheets_hint:
        ws.cell(row=row, column=1, value=hint).font = FONT_BODY
        row += 1

    ws.sheet_view.showGridLines = False
    _autofit(ws)


def _score_fill(value) -> PatternFill | None:
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if v >= 75:
        return FILL_CRITICAL
    if v >= 55:
        return FILL_WARN
    return FILL_OK


def _status_fill(value) -> PatternFill | None:
    text = str(value or "").strip().lower()
    if text == "решено":
        return FILL_OK
    if text == "нерешено":
        return FILL_WARN
    return None


def write_data_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    title: str,
    preferred_cols: list[str],
    score_column: str | None = "score",
    status_column: str | None = None,
) -> tuple[Any, int, int] | None:
    """Возвращает (worksheet, первая строка данных, последняя строка данных) для графиков."""
    prepared = _prepare_df(df, preferred_cols)
    if prepared.empty:
        return None

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    row = _write_title_block(ws, title, f"Записей: {len(prepared)}")
    header_row = row
    for ci, name in enumerate(prepared.columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=name)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 22

    score_idx = None
    status_idx = None
    raw_cols = _pick_columns(df, preferred_cols)
    if score_column and score_column in raw_cols:
        score_idx = raw_cols.index(score_column) + 1
    if status_column:
        status_label = _label(status_column)
        if status_label in prepared.columns:
            status_idx = list(prepared.columns).index(status_label) + 1

    data_start = header_row + 1
    for ri, row_vals in enumerate(prepared.itertuples(index=False), start=data_start):
        fill = FILL_ALT if (ri - data_start) % 2 else None
        status_val = row_vals[status_idx - 1] if status_idx else None
        status_fill = _status_fill(status_val) if status_idx else None
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = FONT_BODY
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if status_fill:
                cell.fill = status_fill
            elif fill:
                cell.fill = fill
            if score_idx and ci == score_idx:
                sf = _score_fill(val)
                if sf:
                    cell.fill = sf
    data_end = data_start + len(prepared) - 1
    ws.freeze_panes = ws.cell(row=data_start, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(prepared.columns))}{data_end}"
    _autofit(ws)
    return ws, data_start, data_end


def _add_top10_charts(wb: Workbook, top10: pd.DataFrame) -> None:
    if top10.empty or "муниципалитет" not in top10.columns:
        return

    df = top10.head(10).copy()
    rows: list[tuple] = []
    for _, row in df.iterrows():
        rows.append(
            (
                str(row["муниципалитет"]),
                float(row["score"]) if "score" in df.columns and pd.notna(row.get("score")) else 0,
                int(row["problem_count"]) if "problem_count" in df.columns and pd.notna(row.get("problem_count")) else 0,
                int(row["total_incidents"]) if "total_incidents" in df.columns and pd.notna(row.get("total_incidents")) else 0,
            )
        )
    if not rows:
        return

    layout = _ensure_charts_layout(
        wb,
        "Top-10 муниципалитетов по индексу проблемности",
    )
    data_start, data_end, col = layout.write_table(
        ["Муниципалитет", "Индекс", "Проблемных", "Всего"],
        rows,
    )
    ws = layout.ws
    n_bars = len(rows)

    if "score" in df.columns:
        chart = BarChart()
        chart.title = "Индекс проблемности (Top-10)"
        _configure_hbar_chart(chart, value_title="Индекс", n_bars=n_bars)
        chart.add_data(_chart_series(ws, col + 1, data_start, data_end), titles_from_data=True)
        chart.set_categories(_chart_categories(ws, col, data_start, data_end))
        layout.add(chart)

    if "problem_count" in df.columns and "total_incidents" in df.columns:
        chart2 = BarChart()
        chart2.title = "Обращения: всего vs проблемных"
        _configure_col_chart(chart2, value_title="Количество", n_bars=n_bars)
        chart2.add_data(_chart_series(ws, col + 2, data_start, data_end), titles_from_data=True)
        chart2.add_data(_chart_series(ws, col + 3, data_start, data_end), titles_from_data=True)
        chart2.set_categories(_chart_categories(ws, col, data_start, data_end))
        layout.add(chart2)


def _add_topics_chart(wb: Workbook, topics_df: pd.DataFrame) -> None:
    if topics_df.empty:
        return
    col_theme = "тема" if "тема" in topics_df.columns else None
    col_count = "count" if "count" in topics_df.columns else None
    if not col_theme or not col_count:
        return

    agg = (
        topics_df.groupby(col_theme, as_index=False)[col_count]
        .sum()
        .sort_values(col_count, ascending=False)
        .head(12)
    )
    if agg.empty:
        return

    layout = _ensure_charts_layout(wb)
    rows = [(str(r[col_theme])[:60], int(r[col_count])) for _, r in agg.iterrows()]
    data_start, data_end, col = layout.write_table(["Тема", "Количество"], rows)
    ws = layout.ws

    chart = BarChart()
    chart.title = "Топ-12 тем (нерешённые)"
    _configure_hbar_chart(chart, value_title="Количество", n_bars=len(rows))
    chart.add_data(_chart_series(ws, col + 1, data_start, data_end), titles_from_data=True)
    chart.set_categories(_chart_categories(ws, col, data_start, data_end))
    layout.add(chart)


def _add_severity_pie(wb: Workbook, breakdown: list[dict]) -> None:
    if not breakdown:
        return
    df = pd.DataFrame(breakdown)
    if df.empty or "count" not in df.columns:
        return
    if "severity" in df.columns:
        df = df[df["severity"].fillna(0).astype(float) > 0]
    if df.empty:
        return
    agg = df.groupby("label" if "label" in df.columns else "severity", as_index=False)["count"].sum()
    if agg.empty:
        return

    layout = _ensure_charts_layout(wb)
    label_col = agg.columns[0]
    rows = [(str(r[label_col]), int(r["count"])) for _, r in agg.iterrows()]
    data_start, data_end, col = layout.write_table(["Класс", "Количество"], rows)
    ws = layout.ws

    pie = PieChart()
    pie.title = "Распределение по классам тяжести (нерешённые)"
    pie.style = 10
    pie.height = 12
    pie.width = 18
    pie.add_data(_chart_series(ws, col + 1, data_start, data_end), titles_from_data=True)
    pie.set_categories(_chart_categories(ws, col, data_start, data_end))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showCatName = True
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.dataLabels.showLegendKey = False
    layout.add(pie)


def write_placeholder_sheet(wb: Workbook, message: str) -> None:
    ws = wb.active
    ws.title = "Сводка"
    _write_title_block(ws, "ZeroProblems — отчёт", message)
    ws.cell(row=4, column=1, value=message).font = FONT_BODY


def build_styled_workbook(payload: ExcelReportPayload) -> Workbook:
    wb = Workbook()
    resolved_stats = payload.resolved_stats or []
    top10 = _enrich_ranking_df(payload.top10, resolved_stats)
    top3 = _enrich_ranking_df(payload.top3, resolved_stats)
    top_all = _enrich_ranking_df(payload.top_all, resolved_stats)
    resolved_df = _resolved_stats_df(resolved_stats)
    labeled_sample = (
        _add_resolution_status(payload.labeled_sample)
        if payload.labeled_sample is not None and not payload.labeled_sample.empty
        else payload.labeled_sample
    )

    has_data = any(
        not df.empty
        for df in (
            payload.top10,
            payload.top3,
            payload.top_all,
            payload.topics_df,
            payload.groups_df,
            payload.reasons_df,
        )
    )
    if not has_data:
        write_placeholder_sheet(wb, payload.summary_text or "Нет данных")
        return wb

    write_summary_sheet(wb, payload)

    if not top10.empty:
        write_data_sheet(
            wb,
            "Top-10",
            top10,
            title="Top-10 проблемных муниципалитетов",
            preferred_cols=RANKING_PREFERRED,
        )
        _add_top10_charts(wb, top10)

    if not top3.empty:
        write_data_sheet(
            wb,
            "Top-3",
            top3,
            title="Top-3 критичных муниципалитета",
            preferred_cols=RANKING_PREFERRED,
        )

    if payload.include_all_municipalities and not top_all.empty:
        write_data_sheet(
            wb,
            "Все МО",
            top_all,
            title="Полный рейтинг муниципалитетов",
            preferred_cols=RANKING_PREFERRED,
        )

    if not resolved_df.empty:
        write_data_sheet(
            wb,
            "Решённость",
            resolved_df,
            title="Решённость проблемных обращений по МО",
            preferred_cols=RESOLVED_STATS_PREFERRED,
            score_column=None,
        )

    if not payload.reasons_df.empty:
        write_data_sheet(
            wb,
            "Причины",
            payload.reasons_df,
            title="Ключевые причины и проблемы",
            preferred_cols=REASONS_PREFERRED,
            score_column=None,
        )

    if not payload.topics_df.empty:
        write_data_sheet(
            wb,
            "Темы",
            payload.topics_df,
            title="Распределение по темам (нерешённые)",
            preferred_cols=TOPIC_PREFERRED,
            score_column=None,
        )
        _add_topics_chart(wb, payload.topics_df)

    if not payload.groups_df.empty:
        write_data_sheet(
            wb,
            "Группы",
            payload.groups_df,
            title="Распределение по группам тем (нерешённые)",
            preferred_cols=GROUP_PREFERRED,
            score_column=None,
        )

    if payload.severity_breakdown:
        sev_df = pd.DataFrame(payload.severity_breakdown)
        write_data_sheet(
            wb,
            "Классы",
            sev_df,
            title="Разбивка по классам тяжести (нерешённые)",
            preferred_cols=["муниципалитет", "severity", "label", "count"],
            score_column=None,
        )
        _add_severity_pie(wb, payload.severity_breakdown)

    if payload.muni_summaries is not None and not payload.muni_summaries.empty:
        write_data_sheet(
            wb,
            "Справки Top-10",
            payload.muni_summaries,
            title="Аналитические справки (Top-10)",
            preferred_cols=["муниципалитет", "summary"],
            score_column=None,
        )

    if payload.top3_summaries is not None and not payload.top3_summaries.empty:
        write_data_sheet(
            wb,
            "Справки Top-3",
            payload.top3_summaries,
            title="Аналитические справки (Top-3)",
            preferred_cols=["муниципалитет", "summary"],
            score_column=None,
        )

    if payload.include_labeled_sample and labeled_sample is not None and not labeled_sample.empty:
        sample_cols = list(labeled_sample.columns)
        if "статус" in sample_cols:
            sample_cols = ["статус"] + [c for c in sample_cols if c != "статус"]
        write_data_sheet(
            wb,
            "Примеры",
            labeled_sample.head(5000),
            title="Размеченные обращения (выборка)",
            preferred_cols=sample_cols,
            score_column=None,
            status_column="статус",
        )

    if payload.meta_rows:
        ws = wb.create_sheet("Служебное")
        ws.append(["Параметр", "Значение"])
        for k, v in payload.meta_rows:
            ws.append([k, v])
        _style_range(ws, 1, 1, 2, fill=FILL_HEADER, font=FONT_HEADER)
        _autofit(ws)

    return wb
